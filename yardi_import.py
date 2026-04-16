"""
Yardi Expense Distribution Report Parser
==========================================
Parses "Expense Distribution (Paid Only)" .xlsx exports from Yardi
and converts them into BoardIQ vendor_data format for benchmarking.

Classifies rows by Account Name (column 1) — a keyword match against the
account's human-readable name — rather than by GL code. This makes the
parser work across different Yardi tenants' charts of accounts
(Century Management, Madison Realty Capital, etc.), which use different
GL numbers for the same types of expenses.

Emits one vendor entry per (category, payee) combination so the building
page shows every vendor, not just one per category. Expenses that don't
match any known category fall into an "OTHER" bucket so they still
appear on the page even without a peer benchmark.
"""

import re
from collections import defaultdict
import openpyxl


# Account Name keyword → BoardIQ category.
# Order matters: the first keyword substring that matches wins, so more
# specific phrases come before generic ones (e.g. "property management
# fee" before "management fee", "water & sewer" before "water").
ACCOUNT_NAME_KEYWORDS = [
    ("water & sewer", "UTILITIES_WATER"),
    ("water and sewer", "UTILITIES_WATER"),
    ("water/sewer", "UTILITIES_WATER"),
    ("water treatment", "WATER_TREATMENT"),
    ("sprinkler", "FIRE_SAFETY"),
    ("fire safety", "FIRE_SAFETY"),
    ("fire alarm", "FIRE_SAFETY"),
    ("elevator", "ELEVATOR_MAINTENANCE"),
    ("janitorial", "CLEANING"),
    ("cleaning", "CLEANING"),
    ("porter", "CLEANING"),
    ("exterminator", "EXTERMINATING"),
    ("pest control", "EXTERMINATING"),
    ("hvac", "HVAC_MAINTENANCE"),
    ("boiler", "HVAC_MAINTENANCE"),
    ("heating", "HVAC_MAINTENANCE"),
    ("chiller", "HVAC_MAINTENANCE"),
    ("workers comp", "INSURANCE"),
    ("workers compensation", "INSURANCE"),
    ("liability", "INSURANCE"),
    ("insurance", "INSURANCE"),
    ("property management fee", "MANAGEMENT_FEE"),
    ("property management", "MANAGEMENT_FEE"),
    ("management fee", "MANAGEMENT_FEE"),
    ("landscap", "LANDSCAPING"),
    ("grounds keeping", "LANDSCAPING"),
    ("plumbing", "PLUMBING_REPAIRS"),
    ("painting", "PAINTING"),
    ("plastering", "PAINTING"),
    ("uniformed security", "SECURITY"),
    ("security staff", "SECURITY"),
    ("doorman", "SECURITY"),
    ("concierge", "SECURITY"),
    ("waste removal", "WASTE_REMOVAL"),
    ("trash", "WASTE_REMOVAL"),
    ("refuse collection", "WASTE_REMOVAL"),
    ("garbage", "WASTE_REMOVAL"),
    ("gas", "UTILITIES_GAS"),
    ("electric", "UTILITIES_ELECTRIC"),
    ("legal", "LEGAL"),
]


# Legacy GL-code fallback, kept as a secondary signal when the account
# name doesn't match any keyword. Values here come from Century
# Management's chart of accounts and are only used as a backstop.
GL_CATEGORY_FALLBACK = {
    "5812": "ELEVATOR_MAINTENANCE",
    "5831": "EXTERMINATING",
    "5803": "CLEANING",
    "5852": "WATER_TREATMENT",
    "5810": "HVAC_MAINTENANCE",
    "5806": "HVAC_MAINTENANCE",
    "5828": "WASTE_REMOVAL",
    "5865": "LANDSCAPING",
    "5870": "LANDSCAPING",
    "6510": "MANAGEMENT_FEE",
    "5837": "SECURITY",
}


def _classify(account_name, gl_code):
    """Classify a GL account into a BoardIQ category.

    Prefers keyword match on the account name (chart-of-accounts agnostic).
    Falls back to a legacy GL-code mapping, then to "OTHER".
    """
    name = (account_name or "").lower()
    for kw, cat in ACCOUNT_NAME_KEYWORDS:
        if kw in name:
            return cat
    if gl_code and gl_code in GL_CATEGORY_FALLBACK:
        return GL_CATEGORY_FALLBACK[gl_code]
    return "OTHER"


def parse_yardi_expense_report(filepath):
    """Parse a single-property Yardi Expense Distribution (Paid Only) .xlsx file.

    Returns:
        dict with keys property_code, period, vendors
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError("Empty spreadsheet")

    return _parse_property_section(rows, 0)


def parse_multi_building_report(filepath):
    """Parse a Yardi export that may contain multiple property sections.

    Each property section starts with an "Expense Distribution (Paid Only)"
    header row in column 0.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError("Empty spreadsheet")

    section_starts = []
    for i, row in enumerate(rows):
        cell0 = str(row[0]).strip() if row[0] else ""
        if "expense distribution" in cell0.lower():
            section_starts.append(i)

    if not section_starts:
        return [_parse_property_section(rows, 0)]

    results = []
    for idx, start in enumerate(section_starts):
        end = section_starts[idx + 1] if idx + 1 < len(section_starts) else len(rows)
        section_rows = rows[start:end]
        try:
            result = _parse_property_section(section_rows, 0)
            results.append(result)
        except Exception as e:
            print(f"[YardiImport] Skipping section at row {start}: {e}")

    return results if results else [_parse_property_section(rows, 0)]


def _parse_property_section(rows, start_offset):
    """Parse one property section from the Yardi report.

    Expected layout (0-indexed within the section):
      Row 0: "Expense Distribution (Paid Only)"
      Row 1: property code (e.g. "148" or ".9504")
      Row 2: "Period: From 01/2025 to 12/2025"
      Row 3: column headers
      Row 4+: GL header rows and invoice rows

    Returns:
        dict with property_code, period, vendors[list of dict]
    """
    property_code_raw = str(rows[1][0]).strip() if len(rows) > 1 and rows[1][0] else ""
    # Some charts prefix the code with a dot (e.g. ".9504") — strip it.
    property_code = property_code_raw.lstrip(".")

    period_raw = str(rows[2][0]).strip() if len(rows) > 2 and rows[2][0] else ""
    period = period_raw.replace("Period: ", "").strip() if period_raw else ""

    # (category, payee_name) -> aggregated total
    totals = defaultdict(float)

    current_gl = None
    current_account_name = None

    for i in range(4, len(rows)):
        row = rows[i]
        if not row or all(c is None for c in row):
            continue

        col0 = str(row[0]).strip() if row[0] else ""
        col1 = str(row[1]).strip() if row[1] and len(row) > 1 else ""

        # Total/grand total rows mark the end of a GL section
        if col0.lower().startswith("total") or col0.lower().startswith("grand total"):
            current_gl = None
            current_account_name = None
            continue

        # GL header row: "5415-0000" in col 0, account name in col 1
        gl_match = re.match(r'^(\d{4})', col0)
        if gl_match:
            current_gl = gl_match.group(1)
            current_account_name = col1 or None
            continue

        # Invoice row: payee in col 3, amount in col 11
        payee_name = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        amount = 0
        if len(row) > 11 and row[11] is not None:
            try:
                amount = float(row[11])
            except (ValueError, TypeError):
                continue

        if not payee_name or amount == 0:
            continue

        category = _classify(current_account_name, current_gl)
        totals[(category, payee_name)] += amount

    # Emit one entry per (category, payee), sorted by category then by
    # spend desc so the biggest vendor in each bucket shows first.
    vendors = []
    for (category, payee), total in sorted(
        totals.items(), key=lambda kv: (kv[0][0], -kv[1])
    ):
        vendors.append({
            "vendor": payee,
            "category": category,
            "annual": round(total),
            "per_unit": 0,  # route fills this once units is known
            "last_bid_year": None,
            "months_left": None,
        })

    return {
        "property_code": property_code,
        "period": period,
        "vendors": vendors,
    }
